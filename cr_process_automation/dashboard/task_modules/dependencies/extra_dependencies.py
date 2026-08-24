import pandas as pd
from threading import Thread
from django.conf import settings
from dashboard.models import CRWiseStatus
from rest_framework.exceptions import APIException
from rest_framework import status
from django_pandas.io import read_frame
from django.utils import timezone
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class CustomThread(Thread):
    def __init__(
        self, group=None, target=None, name=None, args=(), kwargs={}, Verbose=None
    ) -> None:
        Thread.__init__(self, group, target, name, args, kwargs)
        self._returnvalue=None

    def run(self):
        if self._target is not None:
            self._returnvalue = self._target(*self._args, **self._kwargs)

    def join(self):
        Thread.join(self)
        return self._returnvalue


def cr_wise_status_df_maker(date_: datetime) -> pd.DataFrame:
    cr_wise_status_df = read_frame( 
            CRWiseStatus.objects.filter(execution_date=date_+timedelta(days=1), is_active=True).values(*settings.CR_WISE_STATUS_FIELDS)
        )
    return cr_wise_status_df


def _ts_or_none(value):
    if pd.isna(value):
        return None
    ts = pd.Timestamp(value)
    return ts.to_pydatetime()  # tz-aware datetime, cleanly accepted by Django


def workbook_styling(workbook: str):
    wb = load_workbook(workbook)
    if "Planning_Sheet" in wb.sheetnames:
        ws = wb["Planning_Sheet"]
    
    elif "Planning Sheet" in wb.sheetnames:
        ws = wb["Planning Sheet"]
        
    mail_id_sheet_name = ""
    
    if "Mail Id" in wb.sheetnames:
        mail_id_sheet = wb["Mail Id"]
        mail_id_sheet_name = "Mail Id"
    
    elif "Mail_Id" in wb.sheetnames:
        mail_id_sheet = wb["Mail_Id"]
        mail_id_sheet_name = "Mail_Id"
    
    if mail_id_sheet.sheet_state != 'hidden':
        mail_id_sheet.sheet_state = 'hidden'
    
    planning_status_list = ["Planned", "Discussed", "Swapped"]

    # if "Mail Id" in wb.sheetnames:
    #     cs_core_team_list =

    # cs_core_team_list = ["Mandeep", "Parveen", "Rahul"]
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )  # Blue
    header_font = Font(color="FFFFFF", bold=True)  # White bold text

    for cell in ws[1]:  # Row 1 (headers)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = header_font

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align

    for col_num, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                continue

        # Set width with 2 pixels padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
        ws.column_dimensions[column_letter].width = adjusted_width

    # ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    ws.sheet_view.zoomScale = 80

    helper_col_letter_1 = "ZZ"

    range_str_1 = (
        f"'{mail_id_sheet_name}'!$J$2:$J${get_max_row_in_column(mail_id_sheet, column_index_from_string('J'))}"
    )
    range_str_2 =(
        f"'{mail_id_sheet_name}'!$M$2:$M${get_max_row_in_column(mail_id_sheet, column_index_from_string('M'))}"
    )
    range_str_3 = (
        f"'{mail_id_sheet_name}'!$P$2:$P${get_max_row_in_column(mail_id_sheet, column_index_from_string('P'))}"
    )
    
    
    
    dv_1 = DataValidation(
        type="list",
        formula1=range_str_1,
        allow_blank=True,
        error="Select from list only",
        errorTitle="Invalid Executor",
    )
    dv_2 = DataValidation(
        type="list",
        formula1=range_str_2,
        allow_blank=True,
        error="Select from list only",
        errorTitle="Invalid KPI SPOC",
    )
    dv_3 = DataValidation(
        type="list",
        formula1=range_str_3,
        allow_blank = True,
        error="Select from list only",
        errorTitle="Invalid Activity Status"
    )
    ws.add_data_validation(dv_1)
    ws.add_data_validation(dv_2)
    ws.add_data_validation(dv_3)

    max_row = ws.max_row
    dv_1.add(f"O2:O{max_row}")
    dv_2.add(f"Y2:Y{max_row}")
    dv_2.add(f"X2:X{max_row}")
    dv_3.add(f"Q2:Q{max_row}")
    
    formula_string_for_planning_status = f'\"{", ".join(planning_status_list)}\"'

    dv_4 = DataValidation(
        type="list",
        formula1=formula_string_for_planning_status,
        allow_blank=True,
        error="Select from list only",
        errorTitle="Invalid Status",
    )
    ws.add_data_validation(dv_4)

    max_row = ws.max_row
    dv_4.add(f"N2:N{max_row}")
    
    dv5 = DataValidation(
        type="list",
        formula1='\"Yes, No\"',
        allow_blank= True,
        error="Select from list only",
        errorTitle="Invalid Input"
    )
    ws.add_data_validation(dv5)
    
    dv5.add(f"AK2:AK{max_row}")
    
    
    
    # node_type_list = f"'{mail_id_sheet_name}'!$P$2:$P${get_max_row_in_column(mail_id_sheet, column_index_from_string('P'))}"
    
    # dv6= DataValidation(
    #     type="list",
    #     formula1=node_type_list,
    #     allow_blank=True,
    #     error="Select from the list only",
    #     errorTitle="Invalid Input"
    #     )
    # ws.add_data_validation(dv6)
    
    # dv6.add(f"AL2:AL{max_row}")
    # ws.column_dimensions[helper_col_letter_2].hidden = True

    wb.save(workbook)
    wb.close()