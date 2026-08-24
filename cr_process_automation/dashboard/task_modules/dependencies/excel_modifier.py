from ast import Tuple
import os
from numpy._core.numerictypes import str_
import pandas as pd
from functools import singledispatchmethod
from typing import AnyStr, List, Optional, Dict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string


class ExcelModifier:
    def __init__(
        self,
        workbook_path,
        sheet_name: str,
        dataframe: Optional[pd.DataFrame] = None,
        index_required: Optional[bool] = False,
        wrap_text: Optional[bool] = False,
    ):
        self.workbook = None
        self.workbook_to_be_saved = workbook_path

        if dataframe is not None:
            excel_writer = None
            if os.path.exists(str(workbook_path)):
                excel_writer = pd.ExcelWriter(
                    str(workbook_path),
                    engine="openpyxl",
                    mode="a",
                    if_sheet_exists="replace",
                )

            else:
                excel_writer = pd.ExcelWriter(
                    str(workbook_path),
                    engine="openpyxl",
                    mode="w",
                )

            if excel_writer:
                dataframe.to_excel(
                    excel_writer, sheet_name=sheet_name, index=index_required
                )
                # excel_writer.close()
                excel_writer.close()
                del excel_writer

        self.worksheet = None
        self.columns = None
        self.rows = None
        self.header_row = None
        self.header_column = None
        self.headers = None
        self.all_values = None

        self.side = Side(style="medium", color="000000")
        self.border = Border(
            left=self.side, right=self.side, top=self.side, bottom=self.side
        )
        self.fill = None
        self.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=wrap_text
        )
        self.header_font = Font(
            name="Ericsson Hilda", bold=True, size=14, color="FFFFFF"
        )
        self.normal_font = Font(name="Ericsson Hilda", size=11)

        if os.path.exists(workbook_path):
            self.workbook = load_workbook(workbook_path, read_only=False)
            self.worksheet = self.workbook[sheet_name]
            self.columns = self.worksheet.max_column
            self.rows = self.worksheet.max_row
            self.header_row, self.header_column = self.first_row_finder_for_header()

    def first_row_finder_for_header(self):
        i = 1
        j = 1
        breaker = False
        while i <= self.rows:
            while j <= self.columns:
                if self.worksheet.cell(row=i, column=j).value is not None:
                    header_row = i
                    header_column = j
                    breaker = True
                    break
                j += 1
            if breaker:
                break
            i += 1
        return header_row, header_column

    def column_width_adjuster(self):
        col_width = []
        default_font_size = 11

        i = self.header_row
        # print(f"{self.rows =}")

        total_columns = self.columns - self.header_column + 1

        while i <= self.rows:
            # print(f"{i = }")
            j = self.header_column
            while j <= self.columns:
                # print(f"{j = }")
                # print(f"{str(self.worksheet.cell(row=i, column=j).value) = }")

                cell_content_size = len(str(self.worksheet.cell(row=i, column=j).value).strip())
                cell = self.worksheet.cell(row=i, column=j)
                current_font_size = cell.font.size if cell.font and cell.font.size else default_font_size
                scaling_factor = current_font_size / default_font_size
                bold_multiplier = 1.45 if cell.font and cell.font.bold else 1.3
                required_size = (cell_content_size * scaling_factor * bold_multiplier)
                # print(f"{required_size =}")

                if len(col_width) < total_columns:

                    col_width.append(
                        required_size
                    )
                else:
                    list_index_to_be_updated = self.columns - (total_columns + j)
                    # print(f"{list_index_to_be_updated =}")
                    col_width[list_index_to_be_updated] = min(
                        max(col_width[list_index_to_be_updated], required_size), 50
                    )
                    # print(f"{col_width[list_index_to_be_updated] =}")

                j += 1
            i += 1

        i = 0
        while i < len(col_width):
            if col_width[i] < 50:
                col_width[i] += 3
            i += 1

        j = self.header_column
        while j <= self.columns:
            self.worksheet.column_dimensions[get_column_letter(j)].width = col_width[
                j - self.header_column
            ]
            j += 1
    
    def new_column_maker(self, column_name: str, column_index: int|None = None):
        if column_index is None:
            column_index = self.columns + 1
        
        # self.worksheet.insert_cols(column_index)
        if self.headers is None:
            _ = self.get_headers()
        
        if column_name not in self.headers:
            self.worksheet.cell(row=self.header_row, column=column_index).value = column_name
            self.columns += 1
            _ = self.get_headers()
    
    @singledispatchmethod
    def new_row_maker(self, values: pd.Series | List[AnyStr] | List[List[AnyStr]] | pd.DataFrame):
        self.worksheet.append(values)
        self.rows += 1

    @new_row_maker.register(list)
    def _(self, values: list):
        if values and isinstance(values[0], (list, tuple)):
            # This is a list of rows
            for value in values:
                self.worksheet.append(value)
                self.rows += 1
        else:
            # This is a single row
            self.worksheet.append(values)
            self.rows += 1

    @new_row_maker.register(pd.Series)
    def _(self, values: pd.Series):
        self.worksheet.append(values.tolist())
        self.rows += 1

    @new_row_maker.register(pd.DataFrame)
    def _(self, values: pd.DataFrame):
        for value in values.values:
            self.worksheet.append(list(value))
            self.rows += 1
    
    @singledispatchmethod
    def value_adder(self, row:int, column:int, value: AnyStr):
        self.worksheet.cell(row=row, column=column).value = value
        
    
    @value_adder.register
    def _(self, header: str, value: AnyStr, row: int|None = None):
        if self.headers is None:
            _ = self.get_headers()
        
        # print(f"{(str(header).strip() not in self.headers) = }")
        if str(header).strip() not in self.headers:
            # print(self.headers)
            self.new_column_maker(str(header).strip())
        
        if row is None:
            i = 0
            while i < self.columns:
                if str(self.worksheet.cell(row=self.header_row, column=self.header_column + i).value).strip() == str(header).strip():
                    self.worksheet.cell(row=self.rows + 1, column=self.header_column + i).value = value
                    self.rows += 1
                    break
                i += 1
        else:
            i = 0
            while i < self.columns:
                if str(self.worksheet.cell(row=self.header_row, column=self.header_column + i).value).strip() == str(header).strip():
                    self.worksheet.cell(row=row, column=self.header_column + i).value = value
                    break
                i += 1
            
    def get_headers(self):
        self.headers = list(self.worksheet.iter_rows(
                min_row=self.header_row, 
                max_row=self.header_row, 
                min_col=self.header_column, 
                max_col=self.columns,
                values_only=True
            ))[0]
        
        return self.headers
            
    @singledispatchmethod
    def get_data(self, row: int, column: int):
        # print(self.worksheet.title)
        # print(self.worksheet.cell(row=self.header_row, column=column).value)
        # print(self.worksheet.cell(row=row, column=column).value)
        return self.worksheet.cell(row=row, column=column).value

    @get_data.register
    def _(self, row: int, header: AnyStr):
        i = 0
        while i < self.columns:
            header_cell_value = self.worksheet.cell(row=self.header_row, column=self.header_column + i).value
            # print(self.worksheet.title)
            # print(f"{header_cell_value = }")
            
            if header_cell_value is not None and str(header_cell_value).strip() == str(header).strip():
                return self.worksheet.cell(row=row, column=self.header_column + i).value
            i += 1
        return None
    
    # def add_column(self, header: str):
    #     self.workbook_to_be_saved.cell(self.header_row, self.columns + 1).value = header
    #     self.columns += 1
    
    def get_row_based_on_value(self, header: str, value: str) -> int|None:
        value = str(value)
        i = self.header_column
        while i <= self.columns:
            if str(self.worksheet.cell(self.header_row, i).value).strip() == header:
                j = self.header_row + 1
                while j <= self.rows:
                    if str(self.worksheet.cell(j, i).value).strip() == value:
                        return j
                    j += 1
            i += 1
        return None
    
    
    def get_cell_based_on_value(self, header: str|int, value: str) -> Tuple[int|None, int|None]:
        value = str(value).strip()
        if isinstance(header, str):
            header = self.column_index(header)
        
        if header is None:
            return None, None
            
        i = self.header_row + 1
        while i <= self.rows:
            cell_value = self.worksheet.cell(i, header).value
            if cell_value is not None and str(cell_value).strip() == value:
                return i, header
            i +=1
        
        return None, None

        
    def normal_styler(self, not_to_save: bool = False):
        self.fill = PatternFill(
            start_color="3333FF", end_color="3333FF", fill_type="solid"
        )
        i = self.header_column
        while i <= self.columns:
            self.worksheet.cell(row=self.header_row, column=i).fill = self.fill
            self.worksheet.cell(
                row=self.header_row, column=i
            ).alignment = self.alignment
            self.worksheet.cell(row=self.header_row, column=i).font = self.header_font
            self.worksheet.cell(row=self.header_row, column=i).border = self.border
            i += 1

        i = self.header_row + 1
        while i <= self.rows:
            j = self.header_column
            while j <= self.columns:
                self.worksheet.cell(row=i, column=j).alignment = self.alignment
                self.worksheet.cell(row=i, column=j).border = self.border
                self.worksheet.cell(row=i, column=j).font = self.normal_font
                j += 1
            i += 1

        self.column_width_adjuster()
        
        if not not_to_save:
            self.save()
    
    def to_df(self, not_to_save: bool = False) -> pd.DataFrame:
        df = None
        data = list(self.worksheet.iter_rows(
            min_row=self.header_row,
            max_row=self.rows,
            min_col=self.header_column,
            max_col=self.columns,
            values_only=True
        ))
        if not data:
            df = pd.DataFrame()

        # 2. Cleanup: If there are rows that contain only None (but were 'touched' in Excel)
        while data and all(cell is None for cell in data[0]):
            data.pop(0)

        # 3. Create DataFrame
        # Use the first identified non-empty row as headers
        if len(data) > 1:
            df = pd.DataFrame(data[1:], columns=data[0])
        else:
            df = pd.DataFrame(columns=data[0]) if data else pd.DataFrame()
            
        if not not_to_save:
            self.save()
        
        return df

    def to_markdown(self) -> str:
        return self.to_df(True).to_markdown()

    
    def get_all_the_values_in_cell(self):
        self.all_values = self.worksheet.iter_rows(
            min_row=self.header_row+1,
            max_row=self.rows,
            min_col=self.header_column,
            max_col=self.columns,
            values_only=True
        )
    
    def get_all_values_in_a_column(self, header: str):
        all_values_in_column = []
        if not self.headers:
            self.get_headers()
        
        if header in self.headers:
            i = self.header_row + 1
            while i <= self.rows:
                all_values_in_column.append(self.worksheet.cell(i, self.column_index(header)).value)
                i += 1
        
        return all_values_in_column
        
    
    def column_index(self, header: str) -> int|None:
        header = header.strip()
        if not self.headers:
            self.headers = self.get_headers()
        if header in self.headers:
            return self.headers.index(header) + self.header_column
        return None
    

    def row_index(self, value: str) -> int|None:
        value=value.strip()
        i = self.header_row + 1
        j = self.header_column
        while i <= self.rows:
            while j <= self.columns:
                if str(self.worksheet.cell(i, j).value).strip() == value:
                    return i
                j += 1
            i +=1
        
        return None

    
    def colorizer_based_on_value_and_header(self, header: str, value: str, color: str):
        header = header.strip()
        value = value.strip()
        color = color.strip()
        
        row, column = self.get_cell_based_on_value(header, value)
        
        if row and column:
            self.worksheet.cell(row=row, column=column).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            
    def colorizer_based_on_cell_value(self, row: int, column: int, color: str):
        if row and column:
            self.worksheet.cell(row=row, column=column).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
    
    def special_styler(self, color_dict: Dict[AnyStr, List[AnyStr]], not_to_save: bool=False):
        header_font = Font(
            name="Ericsson Hilda", bold=True, size=14, color="000000"
        )
        for color, list_of_columns in color_dict.items():
            pattern_fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
            for column in list_of_columns:
                column_index = column_index_from_string(column)
                self.worksheet.cell(row=self.header_row, column=column_index).fill = pattern_fill
                self.worksheet.cell(row=self.header_row, column=column_index).font = header_font
                self.worksheet.cell(row=self.header_row, column=column_index).border = self.border
                self.worksheet.cell(row=self.header_row, column=column_index).alignment = self.alignment
            
            
        i = self.header_row + 1
        while i <= self.rows:
            j = self.header_column
            while j <= self.columns:
                self.worksheet.cell(row=i, column=j).alignment = self.alignment
                self.worksheet.cell(row=i, column=j).border = self.border
                self.worksheet.cell(row=i, column=j).font = self.normal_font
                j += 1
            i += 1

        self.column_width_adjuster()
        
        if not not_to_save:
            self.save()
        

    def merger(self, range: List = None):
        pass

    def save(self):
        if self.workbook:
            self.workbook.save(self.workbook_to_be_saved)
            self.workbook.close()
            del self.workbook
        
    @property
    def total_columns(self):
        return self.columns
    
    @property
    def total_rows(self):
        return self.rows
    
