from django.core.serializers.json import DjangoJSONEncoder
from pathlib import Path
from datetime import datetime
import sqlite3
import sys
import traceback
import threading
from importlib import import_module
import pandas as pd
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, FileResponse, Http404
from django.shortcuts import render, redirect
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views.decorators.http import require_GET, require_POST
from dashboard.models import MasterCRDatabase
from dashboard.views import _common_context
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import datetime, timedelta
from io import BytesIO
from django.utils import timezone
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Add near other imports
MASTER_CR_FIELDS = [
    "id", "sno", "ms_project", "execution_date", "maintenance_window", "cr_no",
    "priority", "risk", "region", "circle", "activity_description", "node_details", "node_count",
    "bpms_cr_yes_no", "planning_status", "activity_executor",
    "auditor_name", "activity_status", "reason_for_rollback_cancel", "technical_validator",
    "service_affecting", "impact", "test_cases", "kpi_name", "kpi_spoc_night",
    "kpi_spoc_morning", "inter_domain_activity", "inter_domain_kpi_required",
    "inter_domain_measuring_kpis", "activity_type", "vendor", "protocol",
    "execution_type", "cli_availability", "team", "scheduled_start_date",
    "scheduled_end_date", "niam_ticket_required", "niam_node_type", "additional_info",
]

EDITABLE_REGION_CR_FIELDS = [
    "node_details",
    "region",
    "circle",
    "node_count",
    "planning_status",
    "activity_executor",
    "auditor_name",
    "activity_status",
    "technical_validator",
    "reason_for_rollback_cancel",
    "test_cases",
    "kpi_spoc_night",
    "kpi_spoc_morning",
    "activity_type",
    "vendor",
    "protocol",
    "execution_type",
    "cli_availability",
    "niam_ticket_required",
    "niam_node_type",
    "additional_info",
]

CR_HISTORY_EXPORT_FIELDS = [
    "ms_project",
    "execution_date",
    "maintenance_window",
    "cr_no",
    "priority",
    "risk",
    "region",
    "circle",
    "activity_description",
    "node_details",
    "node_count",
    "bpms_cr_yes_no",
    "planning_status",
    "activity_executor",
    "auditor_name",
    "activity_status",
    "reason_for_rollback_cancel",
    "technical_validator",
    "service_affecting",
    "impact",
    "test_cases",
    "kpi_name",
    "kpi_spoc_night",
    "kpi_spoc_morning",
    "inter_domain_activity", 
    "inter_domain_kpi_required",
    "inter_domain_measuring_kpis",  
    "activity_type",
    "vendor",
    "protocol",
    "execution_type",
    "cli_availability",
    "team",
    "scheduled_start_date",
    "scheduled_end_date",
    "niam_ticket_required",
    "niam_node_type",
    "additional_info",
]

ALLOWED_REGION_CR_EDIT_ROLES = {"Admin", "Validator", "Night-SPOC"}

@login_required(login_url="login")
def region_crs_view(request):
    ctx = _common_context(request)
    ctx["selected_option"] = "region_crs"
    ctx["field_labels"] = MASTER_CR_FIELDS
    ctx["editable_region_cr_fields"] = EDITABLE_REGION_CR_FIELDS
    ctx["can_edit_region_crs"] = getattr(request.user, "role", "") in ALLOWED_REGION_CR_EDIT_ROLES
    return render(request, "dashboard/region_crs.html", ctx)

@login_required(login_url="login")
def cr_history_view(request):
    ctx = _common_context(request)
    ctx["selected_option"] = "cr_history"
    return render(request, "dashboard/cr_history.html", ctx)

@require_GET
@login_required(login_url="login")
def fetch_region_cr_details(request):
    date_str = request.GET.get("date", "").strip()

    if not date_str:
        return JsonResponse({"ok": False, "message": "Date is required."}, status=400)

    try:
        parsed_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"ok": False, "message": "Invalid date format."}, status=400)

    result = list(
        MasterCRDatabase.objects.filter(execution_date=parsed_date).values(*MASTER_CR_FIELDS)
    )

    for row in result:
        if row.get("execution_date"):
            row["execution_date"] = row["execution_date"].strftime("%d-%m-%Y")

        if row.get("scheduled_start_date"):
            row["scheduled_start_date"] = timezone.localtime(
                row["scheduled_start_date"]
            ).strftime("%d-%m-%Y %H:%M:%S")

        if row.get("scheduled_end_date"):
            row["scheduled_end_date"] = timezone.localtime(
                row["scheduled_end_date"]
            ).strftime("%d-%m-%Y %H:%M:%S")

    return JsonResponse({
        "ok": True,
        "date": date_str,
        "fields": MASTER_CR_FIELDS,
        "rows": result,
    })

@require_POST
@login_required(login_url="login")
def save_region_cr_details(request):
    user_role = getattr(request.user, "role", "")
    if user_role not in ALLOWED_REGION_CR_EDIT_ROLES:
        return JsonResponse({
            "ok": False,
            "message": "You are not authorized to modify Region CR records."
        }, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({
            "ok": False,
            "message": "Invalid JSON payload."
        }, status=400)

    changes = payload.get("changes", [])
    if not isinstance(changes, list) or not changes:
        return JsonResponse({
            "ok": False,
            "message": "No changes were submitted."
        }, status=400)

    updated_rows = []
    errors = []

    with transaction.atomic():
        for item in changes:
            row_id = item.get("id")
            field_values = item.get("fields", {})

            if not row_id:
                errors.append({"id": None, "message": "Missing row id."})
                continue

            if not isinstance(field_values, dict):
                errors.append({"id": row_id, "message": "Invalid fields payload."})
                continue

            invalid_fields = [f for f in field_values.keys() if f not in EDITABLE_REGION_CR_FIELDS]
            if invalid_fields:
                errors.append({
                    "id": row_id,
                    "message": f"Invalid editable fields: {', '.join(invalid_fields)}"
                })
                continue

            try:
                obj = MasterCRDatabase.objects.get(id=row_id)
            except MasterCRDatabase.DoesNotExist:
                errors.append({"id": row_id, "message": "Record not found."})
                continue

            row_updated_fields = []

            for field_name, raw_value in field_values.items():
                value = raw_value

                if isinstance(value, str):
                    value = value.strip()
                    if value.lower() in {"nan", "na", "n/a", "n.a.", "n.a", "none", "null", "nat"}:
                        value = ""

                if field_name == "node_count":
                    if value in ("", None):
                        value = None
                    else:
                        try:
                            value = int(value)
                        except (TypeError, ValueError):
                            errors.append({
                                "id": row_id,
                                "message": "node_count must be a valid integer."
                            })
                            row_updated_fields = []
                            break

                setattr(obj, field_name, value)
                row_updated_fields.append(field_name)

            if not row_updated_fields:
                continue

            obj.save(update_fields=row_updated_fields)
            updated_rows.append({
                "id": obj.id,
                "cr_no": obj.cr_no,
                "updated_fields": row_updated_fields,
            })

    if errors and not updated_rows:
        return JsonResponse({
            "ok": False,
            "message": "No records were updated.",
            "errors": errors,
        }, status=400)

    return JsonResponse({
        "ok": True,
        "message": f"{len(updated_rows)} record(s) updated successfully.",
        "updated_rows": updated_rows,
        "errors": errors,
    })

def _get_history_start_date(range_key):
    today = timezone.localdate()

    if range_key == "1m":
        return today - timedelta(days=30), today
    if range_key == "3m":
        return today - timedelta(days=90), today
    if range_key == "6m":
        return today - timedelta(days=180), today
    if range_key == "12m":
        return today - timedelta(days=365), today

    return None, None


def _get_cr_history_queryset(date_str=None, range_key=None):
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return None, None, None, "Invalid date format."

        qs = MasterCRDatabase.objects.filter(
            execution_date=selected_date
        ).order_by("cr_no")
        return qs, selected_date, selected_date, None

    start_date, end_date = _get_history_start_date(range_key)
    if not start_date:
        return None, None, None, "Invalid history period selected."

    qs = MasterCRDatabase.objects.filter(
        execution_date__isnull=False,
        execution_date__gte=start_date,
        execution_date__lte=end_date,
    ).order_by("-execution_date", "cr_no")
    return qs, start_date, end_date, None

@require_GET
@login_required(login_url="login")
def fetch_cr_history(request):
    date_str = request.GET.get("date", "").strip()
    range_key = request.GET.get("range", "").strip()

    qs, start_date, end_date, err = _get_cr_history_queryset(date_str, range_key)
    if err:
        return JsonResponse({"ok": False, "message": err}, status=400)

    rows = list(qs.values(*CR_HISTORY_EXPORT_FIELDS))

    for row in rows:
        if row.get("execution_date"):
            row["execution_date"] = row["execution_date"].strftime("%d-%m-%Y")

        if row.get("scheduled_start_date"):
            row["scheduled_start_date"] = timezone.localtime(row["scheduled_start_date"]).strftime("%d-%m-%Y %H:%M:%S")

        if row.get("scheduled_end_date"):
            row["scheduled_end_date"] = timezone.localtime(row["scheduled_end_date"]).strftime("%d-%m-%Y %H:%M:%S")

    numbered_rows = []
    for idx, row in enumerate(rows, start=1):
        numbered_row = {"sno": idx}
        numbered_row.update(row)
        numbered_rows.append(numbered_row)

    if date_str:
        message = f"Showing {len(numbered_rows)} record(s) for {start_date.strftime('%d-%m-%Y')}."
    else:
        range_label_map = {
            "1m": "Last Month CR",
            "3m": "Last 3-Months CR",
            "6m": "Last 6-Months CR",
            "12m": "Last Year CR",
        }
        message = (
            f"Showing {len(numbered_rows)} record(s) for "
            f"{range_label_map.get(range_key, 'selected period')} "
            f"from {start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}."
        )

    return JsonResponse({
        "ok": True,
        "rows": numbered_rows,
        "message": message,
    })


@require_GET
@login_required(login_url="login")
def download_cr_history(request):
    date_str = request.GET.get("date", "").strip()
    range_key = request.GET.get("range", "").strip()

    qs, start_date, end_date, err = _get_cr_history_queryset(date_str, range_key)
    if err:
        return JsonResponse({"ok": False, "message": err}, status=400)

    rows = list(qs.values(*CR_HISTORY_EXPORT_FIELDS))

    numbered_rows = []
    for idx, row in enumerate(rows, start=1):
        numbered_row = {"sno": idx}
        numbered_row.update(row)
        numbered_rows.append(numbered_row)

    df = pd.DataFrame(numbered_rows)

    if not df.empty and "execution_date" in df.columns:
        df["execution_date"] = pd.to_datetime(df["execution_date"]).dt.strftime("%d-%m-%Y")

    if not df.empty and "scheduled_start_date" in df.columns:
        df["scheduled_start_date"] = pd.to_datetime(
                                        df["scheduled_start_date"], utc=True, errors="coerce"
                                        ).dt.tz_convert(timezone.get_current_timezone()).dt.strftime("%d-%m-%Y %H:%M:%S")

    if not df.empty and "scheduled_end_date" in df.columns:
        df["scheduled_end_date"] = pd.to_datetime(
                                        df["scheduled_end_date"], utc=True, errors="coerce"
                                        ).dt.tz_convert(timezone.get_current_timezone()).dt.strftime("%d-%m-%Y %H:%M:%S")

    header_label_map = {
        "sno": "S.No",
        "cr_no": "CR No",
        "ms_project": "MS Project",
        "execution_date": "Execution Date",
        "maintenance_window": "Maintenance Window",
        "region": "Region",
        "circle": "Circle",
        "priority": "Priority",
        "risk": "Risk",
        "planning_status": "Planning Status",
        "activity_status": "Activity Status",
        "vendor": "Vendor",
        "team": "Team",
        "activity_description": "Activity Description",
        "node_details": "Node Details",
        "node_count": "Node Count",
        "bpms_cr_yes_no": "BPMS CR (Yes/No)",
        "activity_executor": "Activity Executor",
        "auditor_name": "Auditor Name",
        "reason_for_rollback_cancel": "Reason For Rollback/Cancel",
        "technical_validator": "Technical Validator",
        "service_affecting": "Service Affecting",
        "impact": "Impact",
        "test_cases": "Test Cases",
        "kpi_name": "KPI Name",
        "kpi_spoc_night": "KPI SPOC Night",
        "kpi_spoc_morning": "KPI SPOC Morning",
        "inter_domain_activity": "Inter Domain Activity",
        "inter_domain_kpi_required": "Inter Domain KPI Required",
        "inter_domain_measuring_kpis": "Inter Domain Measuring KPIs",
        "activity_type": "Activity Type",
        "scheduled_start_date": "Scheduled Start Date",
        "scheduled_end_date": "Scheduled End Date",
        "protocol": "Protocol",
        "execution_type": "Execution Type",
        "cli_availability": "CLI Availability",
        "niam_ticket_required": "NIAM Ticket Required",
        "niam_node_type": "NIAM Node Type",
        "additional_info": "Additional Info",
    }

    df = df.rename(columns=header_label_map)

    range_label_map = {
        "1m": "last_month",
        "3m": "last_3_months",
        "6m": "last_6_months",
        "12m": "last_year",
    }

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="CR History")
        worksheet = writer.sheets["CR History"]

        thin_side = Side(style="thin", color="000000")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0A5EA8", end_color="0A5EA8", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        max_col = worksheet.max_column
        max_row = worksheet.max_row

        for col_idx in range(1, max_col + 1):
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = center_align
            header_cell.border = cell_border

        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, max_col + 1):
                body_cell = worksheet.cell(row=row_idx, column=col_idx)
                body_cell.alignment = center_align
                body_cell.border = cell_border

        for col_idx in range(1, max_col + 1):
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            max_length = max(
                len(str(worksheet.cell(row=r, column=col_idx).value or ""))
                for r in range(1, max_row + 1)
            )
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 40)

        worksheet.freeze_panes = "A2"

    output.seek(0)

    if date_str:
        filename = f"Final_Planning_Sheet_{start_date.strftime('%Y%m%d')}.xlsx"
    else:
        filename = f"cr_history_{range_label_map.get(range_key, 'history')}_{timezone.localdate().strftime('%Y%m%d')}.xlsx"

    return FileResponse(
        output,
        as_attachment=True,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )